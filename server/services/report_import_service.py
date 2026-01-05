import re
import json
from openpyxl import load_workbook
from io import BytesIO
from server.models import db, Analysis, User
import uuid

REQUIRED_COLUMNS = ['user_email', 'business_idea']
OPTIONAL_COLUMNS = [
    'industry', 'target_market', 'location', 'budget', 
    'report_type', 'report_language', 'score',
    'tab_overview', 'tab_market', 'tab_business', 
    'tab_technical', 'tab_financial', 'tab_strategy'
]
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

def validate_email(email, existing_users):
    if not email or not isinstance(email, str):
        return False, "User email is required"
    email = email.strip().lower()
    if not EMAIL_REGEX.match(email):
        return False, "Invalid email format"
    if email not in existing_users:
        return False, f"User with email '{email}' does not exist"
    return True, email

def validate_business_idea(idea):
    if not idea:
        return False, "Business idea is required"
    idea = str(idea).strip()
    if len(idea) < 10:
        return False, "Business idea must be at least 10 characters"
    return True, idea

def validate_report_type(report_type):
    if not report_type or report_type == '':
        return True, 'premium'
    report_type = str(report_type).strip().lower()
    if report_type in ['free', 'premium']:
        return True, report_type
    return False, "Report type must be 'free' or 'premium'"

def validate_report_language(lang):
    if not lang or lang == '':
        return True, 'english'
    lang = str(lang).strip().lower()
    if lang in ['en', 'english']:
        return True, 'english'
    if lang in ['ar', 'arabic']:
        return True, 'arabic'
    return False, "Language must be 'english' or 'arabic'"

def validate_score(score_val):
    if score_val is None or score_val == '':
        return True, None
    try:
        score = int(score_val)
        if score < 0 or score > 100:
            return False, "Score must be between 0 and 100"
        return True, score
    except (ValueError, TypeError):
        return False, "Score must be a valid number"

def parse_json_field(value, field_name):
    if not value or value == '':
        return True, None
    if isinstance(value, dict):
        return True, value
    try:
        if isinstance(value, str):
            parsed = json.loads(value)
            return True, parsed
        return True, None
    except json.JSONDecodeError as e:
        return False, f"Invalid JSON in {field_name}: {str(e)}"

def normalize_column_name(col):
    if not col:
        return None
    col = str(col).strip().lower().replace(' ', '_').replace('-', '_')
    column_mappings = {
        'user_email': 'user_email',
        'email': 'user_email',
        'user': 'user_email',
        'business_idea': 'business_idea',
        'idea': 'business_idea',
        'business': 'business_idea',
        'industry': 'industry',
        'sector': 'industry',
        'target_market': 'target_market',
        'market': 'target_market',
        'location': 'location',
        'country': 'location',
        'region': 'location',
        'budget': 'budget',
        'investment': 'budget',
        'report_type': 'report_type',
        'type': 'report_type',
        'report_language': 'report_language',
        'language': 'report_language',
        'lang': 'report_language',
        'score': 'score',
        'rating': 'score',
        'tab_overview': 'tab_overview',
        'overview': 'tab_overview',
        'tab_market': 'tab_market',
        'market_analysis': 'tab_market',
        'tab_business': 'tab_business',
        'business_model': 'tab_business',
        'tab_technical': 'tab_technical',
        'technical': 'tab_technical',
        'tab_financial': 'tab_financial',
        'financial': 'tab_financial',
        'tab_strategy': 'tab_strategy',
        'strategy': 'tab_strategy',
    }
    return column_mappings.get(col, col)

def parse_excel_file(file_data):
    try:
        wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, "Excel file is empty"
        
        header_row = rows[0]
        if not header_row or all(cell is None for cell in header_row):
            return None, "No header row found"
        
        column_map = {}
        for idx, col in enumerate(header_row):
            normalized = normalize_column_name(col)
            if normalized in ALL_COLUMNS:
                column_map[normalized] = idx
        
        missing_required = [col for col in REQUIRED_COLUMNS if col not in column_map]
        if missing_required:
            return None, f"Missing required columns: {', '.join(missing_required)}"
        
        data_rows = rows[1:]
        parsed_rows = []
        
        for row_num, row in enumerate(data_rows, start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            row_data = {'row_number': row_num}
            for col_name, col_idx in column_map.items():
                value = row[col_idx] if col_idx < len(row) else None
                row_data[col_name] = value
            parsed_rows.append(row_data)
        
        wb.close()
        return {'columns': list(column_map.keys()), 'rows': parsed_rows}, None
        
    except Exception as e:
        return None, f"Failed to parse Excel file: {str(e)}"

def validate_rows(rows):
    existing_users = set(
        email.lower() for (email,) in db.session.query(User.email).all()
    )
    
    validated_rows = []
    
    for row in rows:
        row_num = row.get('row_number', 0)
        errors = []
        warnings = []
        validated_data = {}
        
        valid, result = validate_email(row.get('user_email'), existing_users)
        if valid:
            validated_data['user_email'] = result
        else:
            errors.append(result)
        
        valid, result = validate_business_idea(row.get('business_idea'))
        if valid:
            validated_data['business_idea'] = result
        else:
            errors.append(result)
        
        valid, result = validate_report_type(row.get('report_type'))
        if valid:
            validated_data['report_type'] = result
        else:
            errors.append(result)
        
        valid, result = validate_report_language(row.get('report_language'))
        if valid:
            validated_data['report_language'] = result
        else:
            errors.append(result)
        
        valid, result = validate_score(row.get('score'))
        if valid:
            validated_data['score'] = result
        else:
            warnings.append(result)
        
        validated_data['industry'] = str(row.get('industry', '')).strip() if row.get('industry') else None
        validated_data['target_market'] = str(row.get('target_market', '')).strip() if row.get('target_market') else None
        validated_data['location'] = str(row.get('location', '')).strip() if row.get('location') else None
        validated_data['budget'] = str(row.get('budget', '')).strip() if row.get('budget') else None
        
        tab_fields = ['tab_overview', 'tab_market', 'tab_business', 'tab_technical', 'tab_financial', 'tab_strategy']
        for field in tab_fields:
            valid, result = parse_json_field(row.get(field), field)
            if valid:
                validated_data[field] = result
            else:
                warnings.append(result)
                validated_data[field] = None
        
        status = 'valid' if not errors else 'invalid'
        validated_rows.append({
            'row_number': row_num,
            'status': status,
            'errors': errors,
            'warnings': warnings,
            'data': validated_data,
            'original': row
        })
    
    return validated_rows

def import_reports(validated_rows, skip_invalid=True):
    imported = []
    failed = []
    
    rows_to_import = [r for r in validated_rows if r['status'] == 'valid'] if skip_invalid else validated_rows
    
    for row in rows_to_import:
        if row['status'] != 'valid':
            failed.append({
                'row_number': row['row_number'],
                'errors': row['errors'],
                'user_email': row.get('original', {}).get('user_email', 'Unknown')
            })
            continue
        
        try:
            data = row['data']
            
            analysis = Analysis(
                id=str(uuid.uuid4()),
                user_email=data['user_email'],
                business_idea=data['business_idea'],
                industry=data.get('industry'),
                target_market=data.get('target_market'),
                location=data.get('location'),
                budget=data.get('budget'),
                report_type=data.get('report_type', 'premium'),
                report_language=data.get('report_language', 'english'),
                status='completed',
                score=data.get('score'),
                tab_overview=data.get('tab_overview'),
                tab_market=data.get('tab_market'),
                tab_business=data.get('tab_business'),
                tab_technical=data.get('tab_technical'),
                tab_financial=data.get('tab_financial'),
                tab_strategy=data.get('tab_strategy'),
            )
            
            db.session.add(analysis)
            imported.append({
                'row_number': row['row_number'],
                'user_email': data['user_email'],
                'business_idea': data['business_idea'][:50] + '...' if len(data['business_idea']) > 50 else data['business_idea']
            })
            
        except Exception as e:
            failed.append({
                'row_number': row['row_number'],
                'errors': [str(e)],
                'user_email': row.get('original', {}).get('user_email', 'Unknown')
            })
    
    if imported:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return [], [{'row_number': 0, 'errors': [f'Database error: {str(e)}'], 'user_email': 'All rows'}]
    
    return imported, failed

def get_template_columns():
    return {
        'required': REQUIRED_COLUMNS,
        'optional': OPTIONAL_COLUMNS,
        'all': ALL_COLUMNS
    }
