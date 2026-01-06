import re
import bcrypt
from openpyxl import load_workbook
from io import BytesIO
from server.models import db, User, Role
import secrets
import string

REQUIRED_COLUMNS = ['email']
OPTIONAL_COLUMNS = ['password', 'full_name', 'display_name', 'credits', 'role', 'language', 'phone_number', 'country', 'city']
ALL_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

EMAIL_REGEX = re.compile(r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$')

def generate_referral_code():
    chars = string.ascii_uppercase + string.digits
    return ''.join(secrets.choice(chars) for _ in range(8))

def validate_email(email):
    if not email or not isinstance(email, str):
        return False, "Email is required"
    email = email.strip().lower()
    if not EMAIL_REGEX.match(email):
        return False, "Invalid email format"
    return True, email

def validate_password(password):
    if not password or str(password).strip() == '':
        return True, None
    password = str(password).strip()
    if len(password) < 6:
        return False, "Password must be at least 6 characters"
    return True, password

def validate_credits(credits_val):
    if credits_val is None or credits_val == '':
        return True, 0
    try:
        credits = int(credits_val)
        if credits < 0:
            return False, "Credits cannot be negative"
        return True, credits
    except (ValueError, TypeError):
        return False, "Credits must be a valid number"

def validate_role(role_name, roles_map, default_role_id=None):
    if not role_name or role_name == '':
        role_id = roles_map.get('user', default_role_id)
        if role_id is None:
            return False, "Default 'user' role not found in database"
        return True, role_id
    role_name = str(role_name).strip().lower()
    if role_name in roles_map:
        return True, roles_map[role_name]
    return False, f"Invalid role: {role_name}. Available roles: {', '.join(roles_map.keys())}"

def validate_language(lang):
    if not lang or lang == '':
        return True, 'en'
    lang = str(lang).strip().lower()
    if lang in ['en', 'ar', 'english', 'arabic']:
        return True, 'ar' if lang in ['ar', 'arabic'] else 'en'
    return False, "Language must be 'en' or 'ar'"

def normalize_column_name(col):
    if not col:
        return None
    col = str(col).strip().lower().replace(' ', '_').replace('-', '_')
    column_mappings = {
        'email': 'email',
        'email_address': 'email',
        'e_mail': 'email',
        'password': 'password',
        'pass': 'password',
        'pwd': 'password',
        'full_name': 'full_name',
        'fullname': 'full_name',
        'name': 'full_name',
        'display_name': 'display_name',
        'displayname': 'display_name',
        'nickname': 'display_name',
        'credits': 'credits',
        'credit': 'credits',
        'balance': 'credits',
        'role': 'role',
        'user_role': 'role',
        'role_name': 'role',
        'language': 'language',
        'lang': 'language',
        'phone_number': 'phone_number',
        'phone': 'phone_number',
        'mobile': 'phone_number',
        'country': 'country',
        'city': 'city',
    }
    return column_mappings.get(col, col)

def parse_excel_file(file_data):
    try:
        wb = load_workbook(filename=BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return None, "Excel file is empty"
        
        header_row = rows[0]
        if not header_row or all(cell is None for cell in header_row):
            return None, "No header row found"
        
        column_map = {}
        for idx, col in enumerate(header_row):
            normalized = normalize_column_name(col)
            if normalized in ALL_COLUMNS:
                column_map[normalized] = idx
        
        missing_required = [col for col in REQUIRED_COLUMNS if col not in column_map]
        if missing_required:
            return None, f"Missing required columns: {', '.join(missing_required)}"
        
        data_rows = rows[1:]
        parsed_rows = []
        
        for row_num, row in enumerate(data_rows, start=2):
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            row_data = {'row_number': row_num}
            for col_name, col_idx in column_map.items():
                value = row[col_idx] if col_idx < len(row) else None
                row_data[col_name] = value
            parsed_rows.append(row_data)
        
        wb.close()
        return {'columns': list(column_map.keys()), 'rows': parsed_rows}, None
        
    except Exception as e:
        return None, f"Failed to parse Excel file: {str(e)}"

def validate_rows(rows):
    roles = Role.query.all()
    roles_map = {r.name.lower(): r.id for r in roles}
    
    default_role_id = roles_map.get('user')
    if not default_role_id:
        user_role = Role.query.filter_by(name='user').first()
        if user_role:
            default_role_id = user_role.id
            roles_map['user'] = user_role.id
    
    existing_emails = set(
        email.lower() for (email,) in db.session.query(User.email).all()
    )
    
    validated_rows = []
    batch_emails = set()
    
    for row in rows:
        row_num = row.get('row_number', 0)
        errors = []
        warnings = []
        validated_data = {}
        
        valid, result = validate_email(row.get('email'))
        if valid:
            email = result
            validated_data['email'] = email
            
            if email in existing_emails:
                errors.append(f"Email '{email}' already exists in database")
            elif email in batch_emails:
                errors.append(f"Duplicate email '{email}' in import file")
            else:
                batch_emails.add(email)
        else:
            errors.append(result)
        
        valid, result = validate_password(row.get('password'))
        if valid:
            validated_data['password'] = result
        else:
            errors.append(result)
        
        valid, result = validate_credits(row.get('credits'))
        if valid:
            validated_data['credits'] = result
        else:
            errors.append(result)
        
        valid, result = validate_role(row.get('role'), roles_map, default_role_id)
        if valid:
            validated_data['role_id'] = result
        else:
            errors.append(result)
        
        valid, result = validate_language(row.get('language'))
        if valid:
            validated_data['language'] = result
        else:
            warnings.append(result)
            validated_data['language'] = 'en'
        
        validated_data['full_name'] = str(row.get('full_name', '')).strip() if row.get('full_name') else None
        validated_data['display_name'] = str(row.get('display_name', '')).strip() if row.get('display_name') else None
        validated_data['phone_number'] = str(row.get('phone_number', '')).strip() if row.get('phone_number') else None
        validated_data['country'] = str(row.get('country', '')).strip() if row.get('country') else None
        validated_data['city'] = str(row.get('city', '')).strip() if row.get('city') else None
        
        
        status = 'valid' if not errors else 'invalid'
        validated_rows.append({
            'row_number': row_num,
            'status': status,
            'errors': errors,
            'warnings': warnings,
            'data': validated_data,
            'original': row
        })
    
    return validated_rows

def import_users(validated_rows, skip_invalid=True):
    imported = []
    failed = []
    
    rows_to_import = [r for r in validated_rows if r['status'] == 'valid'] if skip_invalid else validated_rows
    
    for row in rows_to_import:
        if row['status'] != 'valid':
            failed.append({
                'row_number': row['row_number'],
                'errors': row['errors'],
                'email': row.get('original', {}).get('email', 'Unknown')
            })
            continue
        
        try:
            data = row['data']
            password_hash = None
            if data.get('password'):
                password_hash = bcrypt.hashpw(
                    data['password'].encode('utf-8'), 
                    bcrypt.gensalt()
                ).decode('utf-8')
            
            referral_code = generate_referral_code()
            while User.query.filter_by(referral_code=referral_code).first():
                referral_code = generate_referral_code()
            
            user = User(
                email=data['email'],
                password_hash=password_hash,
                full_name=data.get('full_name'),
                display_name=data.get('display_name'),
                role_id=data.get('role_id'),
                credits=data.get('credits', 0),
                language=data.get('language', 'en'),
                phone_number=data.get('phone_number'),
                country=data.get('country'),
                city=data.get('city'),
                referral_code=referral_code,
                email_verified=False,
                is_active=False
            )
            
            db.session.add(user)
            imported.append({
                'row_number': row['row_number'],
                'email': data['email'],
                'full_name': data.get('full_name')
            })
            
        except Exception as e:
            failed.append({
                'row_number': row['row_number'],
                'errors': [str(e)],
                'email': row.get('original', {}).get('email', 'Unknown')
            })
    
    if imported:
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return [], [{'row_number': 0, 'errors': [f'Database error: {str(e)}'], 'email': 'All rows'}]
    
    return imported, failed

def get_template_columns():
    return {
        'required': REQUIRED_COLUMNS,
        'optional': OPTIONAL_COLUMNS,
        'all': ALL_COLUMNS
    }
